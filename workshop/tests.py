from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from users.models import User
from workshop.models import (
    Client, Service, ServiceCategory, OrderService, Part, Supplier, Vehicle,
    WorkOrder,
)


class ExportsTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_superuser("admin", password="p")
        cls.receiver = User.objects.create_user("recv", password="p", role=User.Role.RECEIVER)
        cls.mechanic = User.objects.create_user("mech", password="p", role=User.Role.MECHANIC)

        cat = ServiceCategory.objects.create(name="ТО")
        srv = Service.objects.create(category=cat, name="Замена масла", labor_hours=1, price=Decimal("1500"))
        sup = Supplier.objects.create(name="Поставщик")
        part = Part.objects.create(supplier=sup, name="Фильтр", article="F1", stock=10, price=Decimal("500"))
        cl = Client.objects.create(full_name="Иван Иванов", phone="+70000000000")
        veh = Vehicle.objects.create(client=cl, brand="Lada", model="Vesta", plate="X000XX")

        from django.utils import timezone
        cls.order = WorkOrder.objects.create(
            client=cl, vehicle=veh, receiver=cls.receiver,
            status=WorkOrder.Status.ISSUED, is_paid=True,
        )
        OrderService.objects.create(order=cls.order, service=srv, mechanic=cls.mechanic, quantity=1, cost=srv.price)
        from workshop.models import OrderPart
        OrderPart.objects.create(order=cls.order, part=part, quantity=2, price=part.price)
        cls.order.recalculate_total()
        WorkOrder.objects.filter(pk=cls.order.pk).update(closed_at=timezone.now())

    def test_order_pdf_valid(self):
        self.client.force_login(self.admin)
        r = self.client.get(f"/orders/{self.order.pk}/pdf/")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        body = b"".join(r.streaming_content) if r.streaming else r.content
        self.assertEqual(body[:5], b"%PDF-")

    def test_act_pdf_access(self):
        self.client.force_login(self.receiver)
        self.assertEqual(self.client.get(f"/orders/{self.order.pk}/act/").status_code, 200)
        self.client.force_login(self.mechanic)
        self.assertEqual(self.client.get(f"/orders/{self.order.pk}/act/").status_code, 403)

    def test_excel_structure_and_values(self):
        self.client.force_login(self.admin)
        r = self.client.get("/reports/export.xlsx?date_from=2020-01-01&date_to=2030-01-01")
        self.assertEqual(r.status_code, 200)
        body = b"".join(r.streaming_content) if r.streaming else r.content
        wb = load_workbook(BytesIO(body))
        self.assertIn("Сводка", wb.sheetnames)
        self.assertIn("Заказ-наряды", wb.sheetnames)
        self.assertGreaterEqual(len(wb.sheetnames), 7)
        # Выручка в сводке = total заказа (1500 + 1000 = 2500)
        ws = wb["Сводка"]
        revenue = None
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "Выручка за период, ₽":
                revenue = row[1]
        self.assertAlmostEqual(revenue, 2500.0, places=2)

    def test_excel_access_denied_for_non_admin(self):
        self.client.force_login(self.receiver)
        self.assertEqual(self.client.get("/reports/export.xlsx").status_code, 403)

    def test_reports_page_admin_only(self):
        self.client.force_login(self.mechanic)
        self.assertEqual(self.client.get("/reports/").status_code, 403)
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get("/reports/").status_code, 200)
