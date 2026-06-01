"""Генерация печатных форм (PDF, ReportLab) и выгрузка отчётов (Excel, openpyxl).

PDF строится на ReportLab — чистый Python без системных библиотек
(работает на Windows/Linux/macOS «из коробки»). Кириллица — встроенный
шрифт DejaVu Sans (workshop/fonts/).
"""
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone

ACCENT = "#e8853a"
DARK = "#1f2733"

_FONTS_REGISTERED = False


def _register_fonts():
    global _FONTS_REGISTERED
    if _FONTS_REGISTERED:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    fonts_dir = Path(__file__).resolve().parent / "fonts"
    pdfmetrics.registerFont(TTFont("DejaVu", str(fonts_dir / "DejaVuSans.ttf")))
    pdfmetrics.registerFont(
        TTFont("DejaVu-Bold", str(fonts_dir / "DejaVuSans-Bold.ttf"))
    )
    pdfmetrics.registerFontFamily("DejaVu", normal="DejaVu", bold="DejaVu-Bold")
    _FONTS_REGISTERED = True


def money(value):
    """Денежный формат по-русски: 1 500,00 (неразрывный пробел-разделитель)."""
    s = f"{float(value or 0):,.2f}"
    return s.replace(",", "\u00a0").replace(".", ",")


def render_order_pdf(order, services, parts, doc_type, filename):
    """PDF заказ-наряда или акта выполненных работ."""
    _register_fonts()
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    info = getattr(settings, "WORKSHOP_INFO", {})
    accent = colors.HexColor(ACCENT)
    dark = colors.HexColor(DARK)
    grey = colors.HexColor("#777777")
    light = colors.HexColor("#f4f5f7")

    # Стили: межстрочный интервал (leading) задаём явно ≈ 1.3× кегля,
    # иначе у крупных шрифтов строки наезжают друг на друга.
    base = ParagraphStyle("base", fontName="DejaVu", fontSize=9, leading=13)
    org = ParagraphStyle("org", parent=base, fontName="DejaVu-Bold",
                         fontSize=13, leading=16, textColor=dark)
    org_meta = ParagraphStyle("org_meta", parent=base, fontSize=8, leading=11,
                              textColor=grey)
    title = ParagraphStyle("title", parent=base, fontName="DejaVu-Bold",
                           fontSize=15, leading=19, alignment=TA_RIGHT, textColor=dark)
    title_sub = ParagraphStyle("title_sub", parent=base, fontSize=9, leading=12,
                               alignment=TA_RIGHT, textColor=grey)
    h2 = ParagraphStyle("h2", parent=base, fontName="DejaVu-Bold", fontSize=10.5,
                        leading=14, textColor=dark, spaceBefore=10, spaceAfter=5)
    th = ParagraphStyle("th", parent=base, fontName="DejaVu-Bold", fontSize=8.5,
                        leading=11.5, textColor=colors.white)
    th_r = ParagraphStyle("th_r", parent=th, alignment=TA_RIGHT)
    th_c = ParagraphStyle("th_c", parent=th, alignment=TA_CENTER)
    cell = ParagraphStyle("cell", parent=base, fontSize=8.5, leading=12)
    cell_r = ParagraphStyle("cell_r", parent=cell, alignment=TA_RIGHT)
    cell_c = ParagraphStyle("cell_c", parent=cell, alignment=TA_CENTER)
    muted = ParagraphStyle("muted", parent=cell, textColor=grey)
    grand = ParagraphStyle("grand", parent=cell_r, fontName="DejaVu-Bold",
                           fontSize=12, leading=15)
    grand_v = ParagraphStyle("grand_v", parent=grand, textColor=accent)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=16 * mm, rightMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=18 * mm, title=filename,
    )
    story = []
    cw = doc.width

    # --- Шапка ---
    doc_title = "Акт выполненных работ" if doc_type == "act" else "Заказ-наряд"
    left = [
        Paragraph(info.get("name", "Автосервис"), org),
        Paragraph(
            f"{info.get('address', '')}<br/>тел.: {info.get('phone', '')} · "
            f"ИНН {info.get('inn', '')}", org_meta),
    ]
    right = [
        Paragraph(doc_title, title),
        Paragraph(f"№ {order.id} от {order.created_at:%d.%m.%Y}", title_sub),
    ]
    header = Table([[left, right]], colWidths=[cw * 0.6, cw * 0.4])
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -1), 1.5, accent),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story += [header, Spacer(1, 9)]

    # --- Реквизиты ---
    def kv(label, value):
        return [Paragraph(label, muted), Paragraph(str(value or "—"), cell)]

    v = order.vehicle
    meta = Table(
        [
            kv("Клиент:", order.client.full_name) + kv("Телефон:", order.client.phone),
            kv("Автомобиль:", f"{v.brand} {v.model}" + (f", {v.year} г." if v.year else ""))
            + kv("Гос. номер:", v.plate),
            kv("VIN:", v.vin) + kv("Пробег:", f"{order.mileage_in} км" if order.mileage_in else "—"),
            kv("Приёмщик:", order.receiver.full_name) + kv("Статус:", order.get_status_display()),
        ],
        colWidths=[cw * 0.16, cw * 0.34, cw * 0.16, cw * 0.34],
    )
    meta.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(meta)

    if order.complaint:
        story.append(Spacer(1, 6))
        comp = Table([[Paragraph(f"<b>Заявленная неисправность:</b> {order.complaint}", cell)]],
                     colWidths=[cw])
        comp.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), light),
            ("LINEBEFORE", (0, 0), (0, -1), 3, accent),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(comp)

    def table_style(empty_row_index=None):
        cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), dark),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 1), (-1, -1), 0.4, colors.HexColor("#dddddd")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ]
        if empty_row_index is not None:
            cmds.append(("SPAN", (0, empty_row_index), (-1, empty_row_index)))
        return TableStyle(cmds)

    # --- Работы ---
    story.append(Paragraph("Выполненные работы", h2))
    rows = [[Paragraph("№", th_c), Paragraph("Наименование", th),
             Paragraph("Исполнитель", th), Paragraph("Кол-во", th_c),
             Paragraph("Стоимость, ₽", th_r)]]
    services = list(services)
    for i, si in enumerate(services, 1):
        rows.append([Paragraph(str(i), cell_c), Paragraph(si.service.name, cell),
                     Paragraph(si.mechanic.full_name if si.mechanic else "—", cell),
                     Paragraph(str(si.quantity), cell_c), Paragraph(money(si.cost), cell_r)])
    empty_idx = None
    if not services:
        rows.append([Paragraph("Работы не указаны", muted), "", "", "", ""])
        empty_idx = 1
    works = Table(rows, colWidths=[cw * 0.06, cw * 0.44, cw * 0.27, cw * 0.09, cw * 0.14])
    works.setStyle(table_style(empty_idx))
    story.append(works)

    # --- Запчасти ---
    story.append(Paragraph("Запасные части и материалы", h2))
    prows = [[Paragraph("№", th_c), Paragraph("Наименование", th),
              Paragraph("Кол-во", th_c), Paragraph("Цена, ₽", th_r),
              Paragraph("Сумма, ₽", th_r)]]
    parts = list(parts)
    for i, pi in enumerate(parts, 1):
        name = pi.part.name + (f" ({pi.part.article})" if pi.part.article else "")
        prows.append([Paragraph(str(i), cell_c), Paragraph(name, cell),
                      Paragraph(str(pi.quantity), cell_c), Paragraph(money(pi.price), cell_r),
                      Paragraph(money(pi.line_sum), cell_r)])
    empty_idx = None
    if not parts:
        prows.append([Paragraph("Запчасти не указаны", muted), "", "", "", ""])
        empty_idx = 1
    parts_t = Table(prows, colWidths=[cw * 0.06, cw * 0.50, cw * 0.10, cw * 0.17, cw * 0.17])
    parts_t.setStyle(table_style(empty_idx))
    story.append(parts_t)

    # --- Итоги ---
    story.append(Spacer(1, 9))
    paid = "Оплачено" if order.is_paid else "Не оплачено"
    paid_color = colors.HexColor("#198754") if order.is_paid else colors.HexColor("#b02a37")
    totals = Table(
        [
            [Paragraph("Работы:", cell_r), Paragraph(money(order.services_total) + " ₽", cell_r)],
            [Paragraph("Запчасти:", cell_r), Paragraph(money(order.parts_total) + " ₽", cell_r)],
            [Paragraph("ИТОГО к оплате:", grand), Paragraph(money(order.total) + " ₽", grand_v)],
            [Paragraph("Оплата:", cell_r),
             Paragraph(f"<b>{paid}</b>", ParagraphStyle("pd", parent=cell_r, textColor=paid_color))],
        ],
        colWidths=[cw * 0.78, cw * 0.22],
    )
    totals.setStyle(TableStyle([
        ("LINEABOVE", (0, 2), (-1, 2), 1.5, dark),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(totals)

    if doc_type == "act":
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            "Работы выполнены в полном объёме и в установленный срок. "
            "Стороны претензий друг к другу не имеют.", cell))

    # --- Подписи ---
    story.append(Spacer(1, 28))
    sign = Table(
        [[Paragraph(f"Исполнитель / {info.get('name', '')}", muted), "",
          Paragraph(f"Заказчик / {order.client.full_name}", muted)]],
        colWidths=[cw * 0.46, cw * 0.08, cw * 0.46],
    )
    sign.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (0, 0), 0.6, colors.HexColor("#555555")),
        ("LINEABOVE", (2, 0), (2, 0), 0.6, colors.HexColor("#555555")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(sign)

    gen = timezone.localtime().strftime("%d.%m.%Y %H:%M")

    def footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("DejaVu", 7.5)
        canvas.setFillColor(grey)
        canvas.drawCentredString(
            A4[0] / 2, 11 * mm,
            f"Документ сформирован {gen} · информационная система автосервиса")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


# ===================== Excel-выгрузка отчётов =====================
def build_reports_xlsx(data):
    """Сформировать Excel-книгу из данных отчёта (dict из _compute_reports)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    date_from = data["date_from"].isoformat()
    date_to = data["date_to"].isoformat()
    s = data["summary"]

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1F2733")
    head_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center")

    def autofit(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # 1) Сводка
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Отчёт по автосервису"
    ws["A1"].font = title_font
    ws["A2"] = f"Период: {date_from} — {date_to}"
    ws["A3"] = f"Сформирован: {timezone.localtime():%d.%m.%Y %H:%M}"
    rows = [
        ("Показатель", "Значение"),
        ("Выручка за период, ₽", float(s["revenue"])),
        ("в т.ч. работы, ₽", float(s["works_amount"])),
        ("в т.ч. запчасти, ₽", float(s["parts_amount"])),
        ("Выдано заказов", s["orders"]),
        ("Средний чек, ₽", round(float(s["avg"]), 2)),
        ("Нормо-часов выполнено", round(float(s["labor_hours"]), 1)),
        ("Клиентов обслужено", s["clients_served"]),
        ("Автомобилей обслужено", s["cars_served"]),
        ("Новых клиентов за период", s["new_clients"]),
        ("Средний срок ремонта, дней", round(float(s["avg_repair_days"]), 1)),
    ]
    for i, (a, b) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
    style_header(ws, 5, 2)
    autofit(ws, [32, 18])

    # 2) Все заказ-наряды за период (по дате приёма)
    ws_o = wb.create_sheet("Заказ-наряды")
    ws_o.append(["№", "Дата приёма", "Клиент", "Автомобиль", "Гос. номер",
                 "Статус", "Оплачен", "Сумма, ₽"])
    style_header(ws_o, 1, 8)
    for o in data["all_orders"]:
        ws_o.append([
            o.id, o.created_at.strftime("%d.%m.%Y"), o.client.full_name,
            f"{o.vehicle.brand} {o.vehicle.model}", o.vehicle.plate,
            o.get_status_display(), "Да" if o.is_paid else "Нет", float(o.total),
        ])
    autofit(ws_o, [6, 14, 28, 22, 14, 18, 10, 14])

    # 3) Выручка по месяцам
    ws2 = wb.create_sheet("Выручка по месяцам")
    ws2.append(["Месяц", "Выручка, ₽"])
    style_header(ws2, 1, 2)
    for label, value in zip(data["revenue_chart"]["labels"], data["revenue_chart"]["data"]):
        ws2.append([label, float(value)])
    autofit(ws2, [18, 18])

    # 4) Топ клиентов
    ws_c = wb.create_sheet("Топ клиентов")
    ws_c.append(["Клиент", "Заказов", "Сумма, ₽"])
    style_header(ws_c, 1, 3)
    for r in data["top_clients"]:
        ws_c.append([r["client__full_name"], r["cnt"], float(r["revenue"] or 0)])
    autofit(ws_c, [32, 12, 16])

    # 5) Механики
    ws3 = wb.create_sheet("Механики")
    ws3.append(["Механик", "Работ", "Сумма, ₽"])
    style_header(ws3, 1, 3)
    for r in data["mechanic_rows"]:
        ws3.append([r["name"], r["works"], float(r["amount"] or 0)])
    autofit(ws3, [28, 12, 16])

    # 6) Услуги
    ws4 = wb.create_sheet("Услуги")
    ws4.append(["Услуга", "Количество", "Выручка, ₽"])
    style_header(ws4, 1, 3)
    for r in data["service_rows"]:
        ws4.append([r["service__name"], int(r["cnt"]), float(r["revenue"] or 0)])
    autofit(ws4, [36, 14, 16])

    # 7) Запчасти (движение)
    ws5 = wb.create_sheet("Запчасти")
    ws5.append(["Запчасть", "Израсходовано", "Сумма, ₽"])
    style_header(ws5, 1, 3)
    for r in data["parts_rows"]:
        ws5.append([r["part__name"], int(r["qty"]), float(r["amount"] or 0)])
    autofit(ws5, [36, 16, 16])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="report_{date_from}_{date_to}.xlsx"'
    return response
